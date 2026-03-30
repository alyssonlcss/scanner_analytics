"""
Processing pipeline that orchestrates all services.

This module provides the main processing pipeline that coordinates
data loading, calculation, aggregation, and report generation.
"""

from pathlib import Path
from typing import Optional
import pandas as pd
import logging

from ..config import Settings, get_settings
from ..core.models import ProcessingResult
from .data_loader import DataLoaderService
from .calculator import CalculatorService
from .aggregator import AggregatorService
from .excel_formatter import ExcelFormatter, get_excel_formatter

logger = logging.getLogger(__name__)


class ProcessingPipeline:
    """
    Main processing pipeline for displacement analysis.
    
    Orchestrates all services to process displacement data from
    CSV input to calculated metrics and aggregated reports.
    """
    
    def __init__(self, settings: Optional[Settings] = None):
        """
        Initialize the processing pipeline.
        
        Args:
            settings: Application settings. If None, uses default settings.
        """
        self._settings = settings or get_settings()
        
        # Initialize services
        self._loader = DataLoaderService(self._settings)
        self._calculator = CalculatorService(self._settings)
        self._aggregator = AggregatorService(self._settings)
        self._excel_formatter = get_excel_formatter()
        # Apply excel theme from settings if provided via .env
        try:
            theme = getattr(self._settings, "excel_theme", None)
            if theme:
                self._excel_formatter = self._excel_formatter.with_theme(theme)
        except Exception:
            pass
    
    def run(self, input_path: Optional[Path] = None) -> ProcessingResult:
        """
        Execute the full processing pipeline.
        
        Args:
            input_path: Optional path to input file. Uses default if not provided.
            
        Returns:
            ProcessingResult containing all outputs and statistics
        """
        result = ProcessingResult()
        
        try:
            # Step 1: Load data
            logger.info("=" * 60)
            logger.info("STEP 1: Loading data")
            logger.info("=" * 60)
            
            df = self._loader.load(input_path)
            columns = self._loader.resolved_columns
            # Filtrar equipes pelo padrão ABC-YX-
            area_prefixes = ["ACU", "ITJ", "ITK", "TRR"]
            tipo_prefixes = ["SG", "SP", "RD", "TR"]
            valid_prefixes = [f"{a}-{t}-" for a in area_prefixes for t in tipo_prefixes]
            col_equipe = columns.get("equipe")
            if col_equipe and col_equipe in df.columns:
                df = df[df[col_equipe].astype(str).str.startswith(tuple(valid_prefixes))]
            result.total_records = len(df)
            
            # Step 2: Calculate metrics
            logger.info("=" * 60)
            logger.info("STEP 2: Calculating metrics")
            logger.info("=" * 60)
            
            df_calculated = self._calculator.process(df, columns)
            result.df_calculated = df_calculated
            
            # Salva o DataFrame calculado diretamente, sem remover colunas _dt (não existem mais)
            # Exporta apenas CSV do calculado
            import os
            csv_dir = os.path.join(os.path.dirname(self._settings.output_calculated_path), 'csv')
            os.makedirs(csv_dir, exist_ok=True)
            df_calculated.to_csv(os.path.join(csv_dir, 'deslocamento_calculado.csv'), sep=';', index=False, encoding='utf-8')
            
            # Step 3: Filter by status
            logger.info("=" * 60)
            logger.info("STEP 3: Filtering by status")
            logger.info("=" * 60)
            
            df_productive, df_unproductive = self._aggregator.filter_by_status(
                df_calculated, columns
            )
            
            result.productive_records = len(df_productive)
            result.unproductive_records = len(df_unproductive)
            
            # Step 4: Aggregate general averages (Média Geral) using all calculated records
            logger.info("=" * 60)
            logger.info("STEP 4: Aggregating general averages (Média Geral)")
            logger.info("=" * 60)

            # Use the full calculated dataset to produce the general averages
            if not df_calculated.empty:
                result.df_geral_averages = self._aggregator.aggregate(
                    df_calculated, columns, "geral"
                )

                if result.df_geral_averages is not None:
                        # Remove coluna 'Intervalo, Retorno a base' se existir
                        if "Intervalo, Retorno a base" in result.df_geral_averages.columns:
                            result.df_geral_averages.drop(columns=["Intervalo, Retorno a base"], inplace=True)
                        # Exporta apenas CSV das médias gerais (Média Geral)
                        result.df_geral_averages.to_csv(os.path.join(csv_dir, 'medias_por_equipe_dia.csv'), sep=';', index=False, encoding='utf-8')
            
            # Step 5: Aggregate unproductive records
            logger.info("=" * 60)
            logger.info("STEP 5: Aggregating unproductive records")
            logger.info("=" * 60)
            
            if not df_unproductive.empty:
                result.df_unproductive_averages = self._aggregator.aggregate(
                    df_unproductive, columns, "improdutivas"
                )
                
                if result.df_unproductive_averages is not None:
                        # Remove coluna 'Intervalo, Retorno a base' se existir
                        if "Intervalo, Retorno a base" in result.df_unproductive_averages.columns:
                            result.df_unproductive_averages.drop(columns=["Intervalo, Retorno a base"], inplace=True)
                        # Exporta apenas CSV das médias improdutivas
                        result.df_unproductive_averages.to_csv(os.path.join(csv_dir, 'medias_Improdutivas_por_equipe_dia.csv'), sep=';', index=False, encoding='utf-8')
            
            # Calculate team count
            col_equipe = columns.get("equipe")
            if col_equipe and col_equipe in df_calculated.columns:
                result.total_teams = df_calculated[col_equipe].nunique()
            
            result.success = True
            result.message = "Processing completed successfully"
            
        except FileNotFoundError as e:
            result.success = False
            result.message = f"File not found: {e}"
            result.processing_errors.append(str(e))
            logger.error(result.message)
            
        except Exception as e:
            result.success = False
            result.message = f"Processing failed: {e}"
            result.processing_errors.append(str(e))
            logger.exception("Pipeline execution failed")
        
        return result
    
    def _save_dataframe(
        self,
        df: pd.DataFrame,
        path: Path,
        description: str,
        sheet_name: str = "Dados",
        is_aggregated: bool = False
    ) -> None:
        """Save DataFrame to Excel file with formatting and also as CSV."""
        try:
            # Choose theme per sheet/type: prefer specific themes when present
            try:
                title_lower = (sheet_name or "").lower()
                path_lower = (str(path) if path is not None else "").lower()
                theme_key = "default"
                if hasattr(self._settings, "excel_themes"):
                    themes = self._settings.excel_themes
                    # Deslocamento (full calculated records)
                    if "deslocamento" in title_lower or "deslocamento" in path_lower:
                        theme_key = "deslocamento"
                    # Média Geral (all records aggregated)
                    elif is_aggregated and ("geral" in title_lower or "média geral" in title_lower or "media geral" in title_lower):
                        theme_key = "medias_geral"
                    # Medias improdutivas
                    elif is_aggregated and ("improdut" in title_lower or "improdut" in path_lower or "improdutivas" in title_lower or "improdutivas" in path_lower):
                        theme_key = "medias_improdutivas"
                    # Generic medias sheet fallback
                    elif is_aggregated or "media" in title_lower or "média" in title_lower or "medias" in title_lower:
                        theme_key = "medias"

                    theme = themes.get(theme_key, {})
                    if theme:
                        self._excel_formatter.with_theme(theme)
            except Exception:
                pass

            # Use Excel formatter for nice output
            success = self._excel_formatter.export(
                df=df,
                path=path,
                sheet_name=sheet_name,
                summary_identifier="GERAL" if is_aggregated else "",
                freeze_header=True
            )
            if success:
                logger.info(f"{description} saved to: {path}")
            else:
                # Fallback to basic Excel export
                df.to_excel(path, index=False, sheet_name=sheet_name)
                logger.info(f"{description} saved (basic format) to: {path}")

            # Salvar também como CSV
            import os
            csv_dir = os.path.join(os.path.dirname(path), 'csv')
            os.makedirs(csv_dir, exist_ok=True)
            csv_path = os.path.join(csv_dir, os.path.splitext(os.path.basename(path))[0] + '.csv')
            df.to_csv(csv_path, sep=';', index=False, encoding='utf-8')
            logger.info(f"{description} also saved as CSV to: {csv_path}")

        except Exception as e:
            logger.error(f"Failed to save {description}: {e}")
    
    @property
    def loader(self) -> DataLoaderService:
        """Get the data loader service."""
        return self._loader
    
    @property
    def calculator(self) -> CalculatorService:
        """Get the calculator service."""
        return self._calculator
    
    @property
    def aggregator(self) -> AggregatorService:
        """Get the aggregator service."""
        return self._aggregator

    def export_analysis_excel(self, sheets: list, output_path: Path = None) -> None:
        """
        Exporta múltiplos DataFrames em abas de um único arquivo Excel, com formatação.
        sheets: lista de tuplas (nome_aba, dataframe, kwargs_export)
        """
        import os
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.worksheet import Worksheet
        from openpyxl import Workbook

        if not sheets:
            raise ValueError("Nenhuma aba fornecida para exportação.")

        if output_path is None:
            output_path = os.path.join(os.path.dirname(self._settings.output_calculated_path), 'analise_apontamento.xlsx')

        wb = Workbook()
        # Remove sheet padrão criada
        if 'Sheet' in wb.sheetnames:
            std = wb['Sheet']
            wb.remove(std)

        for sheet_name, df, export_kwargs in sheets:
            ws = wb.create_sheet(sheet_name)
            # Usa ExcelFormatter para formatar a aba
            self._excel_formatter.export(
                df=df,
                path=None,  # Não salva, só formata na worksheet
                sheet_name=sheet_name,
                worksheet=ws,
                **(export_kwargs or {})
            )

        wb.save(output_path)
        wb.close()
